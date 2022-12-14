import os
import sys
import matplotlib.pyplot as plt
import matplotlib.dates as md
import winsound
import pandas as pd
import requests
from threading import Thread
from kucoin.client import Client
from matplotlib.ticker import MaxNLocator
from matplotlib.dates import HourLocator, MinuteLocator, SecondLocator
from openpyxl import load_workbook, Workbook
from datetime import datetime
from PyQt6 import uic, QtWidgets
from PyQt6.QtCore import QThread, pyqtSignal, QObject, Qt
from PyQt6.QtGui import QDoubleValidator, QColor, QCloseEvent
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchElementException
from time import sleep

def alarm():
    duration = 500
    freq = 750

    for _ in range(3):
        winsound.Beep(freq, duration)

try:
    wb = load_workbook('Closed Trades.xlsx')
    ws = wb.worksheets[0]
except FileNotFoundError:
    headers_row = ["Pair", "Category", "Open", "Close", "Change % on closure", "Close Condition", "TP", "SL", "Collective"]
    wb = Workbook()
    ws = wb.active
    ws.append(headers_row)

chart_log_wb = load_workbook('Chart Data Log.xlsx')
chart_log_ws = chart_log_wb.worksheets[0]

track = {
    'MT': [],
    'ST': []
}

# Global variables
chart_dir = 'charts'
is_single_sytem_activated = False

if not os.path.exists(chart_dir):
    os.mkdir(chart_dir)
    
class ChartLogger(QObject):
    def __init__(self, st_table):
        super().__init__()

        self.st_table = st_table

    def run(self):
        pairs = []
        column_count = chart_log_ws.max_column

        for index in range(3, column_count+1):
            pairs.append(chart_log_ws.cell(1, index).value)

        while True:
            changes = []

            row_count = self.st_table.rowCount()
            
            for pair in pairs:
                change = ''
                for index in range(row_count):
                    item = self.st_table.item(index, 0)

                    if item:
                        if item.text() == pair:
                            change_item = self.st_table.item(index, 1)

                            if change_item:
                                change = change_item.text()
                            else:
                                change = ''
                                
                            break

                changes.append(change)

            current_datetime = datetime.now()
            current_date = current_datetime.strftime('%Y-%m-%d')
            current_time = current_datetime.strftime('%H:%M:%S')

            row = [current_date, current_time]
            row += changes

            chart_log_ws.append(row)
            chart_log_wb.save('Chart Data Log.xlsx')

            sleep(60)

class Pricer(QObject):
    progress = pyqtSignal(dict)
    finished = pyqtSignal()

    def __init__(self, single_system_table):
        super().__init__()

        self.single_system_table = single_system_table

        api_key = "62de0d8d5ce91e0001765393"
        api_secret = "234a6141-767e-4dd9-ac0f-7b2427c43958"
        api_passphrase = "richardapi"

        self.client = Client(api_key, api_secret, api_passphrase)
        self.result = {}

    def run(self):
        while True:
            row_count = self.single_system_table.rowCount()
            pairs_from_table = []

            for row_index in range(row_count):
                item = self.single_system_table.item(row_index, 0)
                buy = self.single_system_table.item(row_index, 1)
                sell = self.single_system_table.item(row_index, 2)

                if item:
                    if item.text() != "":
                        is_added = False

                        if buy:
                            if buy.background().color() == QColor(0, 255, 0):
                                pairs_from_table.append(item.text().replace('USDT', '3L-USDT'))
                                is_added = True
                        
                        if sell:
                            if sell.background().color() == QColor(255, 0, 0):
                                pairs_from_table.append(item.text().replace('USDT', '3S-USDT'))
                                is_added = True

                        if not is_added:
                            pairs_from_table.append(item.text().replace('USDT', '-USDT'))

            pairs_group = [pairs_from_table[i:i+10] for i in range(0, len(pairs_from_table), 10)]

            self.result = {}

            for group in pairs_group:
                thread_count = len(group)
                threads = []
                for index in range(thread_count):
                    thread = Thread(target=self.getPrice, args=(group[index], ))
                    threads.append(thread)
                    thread.start()

                for thread in threads:
                    thread.join()

            for pair in pairs_from_table:
                for row_index in range(row_count):
                    item = self.single_system_table.item(row_index, 0)

                    if item:
                        if item.text() == pair.replace('-', '').replace('3L', '').replace('3S', ''):
                            try:
                                c_price = float(self.result[pair])
                                c_to_h_price = float(self.single_system_table.item(row_index, 5).text())
                                grid_size = float(self.single_system_table.item(row_index, 6).text())

                                h_price = c_price + (c_to_h_price * c_price)
                                l_price = ((100 - grid_size) * 0.01) * h_price

                                self.single_system_table.setItem(row_index, 4, QtWidgets.QTableWidgetItem('{:.6f}'.format(c_price)))
                                self.single_system_table.setItem(row_index, 7, QtWidgets.QTableWidgetItem('{:.6f}'.format(h_price)))
                                self.single_system_table.setItem(row_index, 8, QtWidgets.QTableWidgetItem('{:.6f}'.format(l_price)))
                            except:
                                pass

                            break

            sleep(0.1)

    def getPrice(self, pair):
        try:
            self.result[pair] = self.client.get_ticker(pair)['price']
        except:
            pass

class Webhook(QObject):
    progress = pyqtSignal(dict)
    finished = pyqtSignal()

    def run(self):
        is_cleaned = False

        while True:
            if is_single_sytem_activated:
                if not is_cleaned:
                    requests.get('http://rick26754.pythonanywhere.com/clear')
                    
                    is_cleaned = True

                r = requests.get('http://rick26754.pythonanywhere.com/')
                
                self.progress.emit(r.json())

                sleep(3)
            else:
                is_cleaned = False

class Bot(QObject):
    progress = pyqtSignal(list, list, bool)
    finished = pyqtSignal()

    close_list = []
    open_list = []

    open_list_buff = []
    close_list_buff = []

    investment = 100

    def __init__(self, single_system_table, single_system_inv):
        super().__init__()
        self.single_system_table = single_system_table
        self.single_system_inv = single_system_inv

    def run(self):
        login_file = open('login.txt', 'r')
        info = login_file.read()
        email = info.split(':')[0].strip()
        password = info.split(':')[1].strip()
        login_file.close()

        login_url = 'https://bitsgap.com/sign-in?d=app'

        service = Service(ChromeDriverManager().install())
        options = Options()
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        driver = webdriver.Chrome(service=service, options=options)
        driver.maximize_window()

        driver.get(login_url)

        # Login
        self.progress.emit(['Signing to Bitsgap...'], [0], False)
        while True:
            try:
                email_input = driver.find_element(By.ID, 'email')
                password_input = driver.find_element(By.ID, 'password')

                email_input.clear()
                password_input.clear()

                ac = ActionChains(driver)
                ac.move_to_element(email_input)
                sleep(0.5)
                ac.click()
                sleep(0.5)
                ac.send_keys(email)
                sleep(2)
                ac.move_to_element(password_input)
                sleep(0.5)
                ac.click()
                sleep(0.5)
                ac.send_keys(password)
                sleep(2)
                ac.perform()
                password_input.submit()
            except StaleElementReferenceException:
                continue
            
            break
        self.progress.emit([f'Signed as {email}'], [0], False)

        # Switch to demo
        self.progress.emit(['Switching to demo...'], [0], False)
        sleep(5)
        driver.get('https://app.bitsgap.com/my-exchanges')
        driver.find_elements(By.CLASS_NAME, 'RX0wjSyh4wgkc8ZhpkRW')[-1].click()

        # Extract table
        sleep(3)

        self.progress.emit(['Switched to demo'], [0], False)

        driver.get('https://app.bitsgap.com/bot')

        self.progress.emit(['Loading finished!'], [0], False)

        while True:
            if len(self.close_list) > 0:
                self.close_list_buff = self.close_list.copy()
                self.close_list = []

                if len(self.open_list) > 0:
                    self.open_list_buff = self.open_list.copy()
                    self.open_list = []

                print('Current close buff', self.close_list_buff)
                print('Current open buff', self.open_list_buff)
                # print("Close list:", self.close_list_buff)

                if len(self.close_list_buff) > 0:
                    self.progress.emit(["Closing pairs..."], [0], False)
                    self.closePair(driver)

                if len(self.open_list_buff) > 0:
                    self.progress.emit(['Opening pairs...'], [0], False)
                    self.openPair(driver)

                self.close_list_buff = []
                self.open_list_buff = []
            
                self.progress.emit(['Extracting change...'], [0], False)
            else:
                self.progress.emit(['Extracting change...'], [0], False)

                is_passed = False

                while not is_passed:
                    try:
                        self.extract(driver)
                        is_passed = True
                    except:
                        is_passed = False

                sleep(10)

    def setOpen(self, pairs):
        self.open_list += pairs
        print('Current open list', self.open_list)

    def setClose(self, pairs):
        self.close_list += pairs
        print('Current close list', self.close_list)

    def openPair(self, driver):
        for index in range(len(self.open_list_buff)):
            pair = self.open_list_buff[index]
            print('Opening', pair)
            pair_for_single_table = pair.replace('/', '').replace('3S', '').replace('3L', '')

            row_count = self.single_system_table.rowCount()

            for row_index in range(row_count):
                item = self.single_system_table.item(row_index, 0)

                if item:
                    if item.text() == pair_for_single_table:
                        pair_for_current_price = pair.replace('/', '-')

                        api_key = "62de0d8d5ce91e0001765393"
                        api_secret = "234a6141-767e-4dd9-ac0f-7b2427c43958"
                        api_passphrase = "richardapi"

                        client = Client(api_key, api_secret, api_passphrase)

                        c_price = float(client.get_ticker(pair_for_current_price)['price'])
                        c_to_h_price = float(self.single_system_table.item(row_index, 5).text())
                        grid_size = float(self.single_system_table.item(row_index, 6).text())

                        h_price = c_price + (c_to_h_price * c_price)
                        l_price = ((100 - grid_size) * 0.01) * h_price

                        low = '{:.6f}'.format(l_price)
                        high = '{:.6f}'.format(h_price)
                        grid = self.single_system_table.item(row_index, 3).text()

                        break

            try:
                self.investment = float(self.single_system_inv.text())
            except:
                pass

            try:
                cross_button = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//button[@class="MuiButtonBase-root MuiIconButton-root MuiIconButton-sizeLarge MJl9JyxF7DXg4KHqMyOg css-1gzpori"]')))
                # cross_button.click()
                driver.execute_script("arguments[0].click();", cross_button)
            except TimeoutException:
                pass

            new_bot_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="bot-start-new-bot-button"]')))
            # new_bot_button.click()
            driver.execute_script("arguments[0].click();", new_bot_button)

            start_bot_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//div[@data-test="start-sbot"]')))
            # start_bot_button.click()
            driver.execute_script("arguments[0].click();", start_bot_button)

            exchange_button = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, '//button[@class="aQXLoSia4k1esjIDAFwW eNHVE8z4gnaP_uzDASJz MuiButton-root MuiButton-text MuiButton-textPrimary MuiButton-sizeMedium MuiButton-textSizeMedium MuiButtonBase-root  css-pev4aq"]')))[0]
            # ac = ActionChains(driver)
            # ac.move_to_element(exchange_button)
            # ac.click()
            # ac.perform()
            # exchange_button.click()
            driver.execute_script("arguments[0].click();", exchange_button)

            search_exchange_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//input[@placeholder="Search by exchanges"]')))
            search_exchange_input.clear()
            search_exchange_input.send_keys('Kucoin')

            found_exchange = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, '//div[@class="MuiTableRow-root Ww5Ht1SCZfv3zC5nQFwh GHjFMeuMOZ1mUTMsHpSZ rSbNWzKWzIJZjdD2kDng css-1gqug66"]')))[-1]
            # found_exchange.click()
            driver.execute_script("arguments[0].click();", found_exchange)
    
            sleep(1)

            select_pair_button = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, '//button[@class="aQXLoSia4k1esjIDAFwW eNHVE8z4gnaP_uzDASJz MuiButton-root MuiButton-text MuiButton-textPrimary MuiButton-sizeMedium MuiButton-textSizeMedium MuiButtonBase-root  css-pev4aq"]')))[-1]
            # select_pair_button.click()
            driver.execute_script("arguments[0].click();", select_pair_button)

            search_pair_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//input[@placeholder="Search by pair"]')))
            search_pair_input.clear()
            search_pair_input.send_keys(pair)

            found_pair = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, '//div[@class="MuiTableRow-root Ww5Ht1SCZfv3zC5nQFwh GHjFMeuMOZ1mUTMsHpSZ rSbNWzKWzIJZjdD2kDng css-1gqug66"]')))[-1]
            # found_pair.click()
            driver.execute_script("arguments[0].click();", found_pair)

            sleep(1)

            input_items = driver.find_elements(By.XPATH, '//input[@class="text-input__input text-input__input_align_left MuiFilledInput-input MuiInputBase-input css-2bxn45"]')

            low_input = input_items[0]
            high_input = input_items[1]
            grid_input = input_items[3]

            low_input.send_keys(Keys.CONTROL, 'a')
            sleep(1)
            low_input.send_keys(Keys.DELETE)
            sleep(1)
            low_input.send_keys(low)

            high_input.send_keys(Keys.CONTROL, 'a')
            sleep(1)
            high_input.send_keys(Keys.DELETE)
            sleep(1)
            high_input.send_keys(high)

            grid_input.send_keys(Keys.CONTROL, 'a')
            sleep(1)
            grid_input.send_keys(Keys.DELETE)
            sleep(1)
            grid_input.send_keys(grid)

            investment_input = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//input[@class="text-input__input text-input__input_align_left MuiFilledInput-input MuiInputBase-input MuiInputBase-inputAdornedEnd css-ftr4jk"]')))
            investment_input.send_keys(Keys.CONTROL, 'a')
            sleep(1)
            investment_input.send_keys(Keys.DELETE)
            sleep(1)
            investment_input.send_keys(str(self.investment))

            # try:
            start_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//button[@data-test="bot-submit-button"]')))
            # start_button.click()
            driver.execute_script("arguments[0].click();", start_button)

            confirm_button = driver.find_element(By.XPATH, '//button[@data-test="bot-preview-confirm-button"]')
            # confirm_button.click()
            driver.execute_script("arguments[0].click();", confirm_button)

            try:
                WebDriverWait(driver, 300).until(lambda s: s.find_element(By.XPATH, '//div[@class="vfOtxORs1c0oUEoFMKdP"]').text == 'Your bot has been started')
            except:
                driver.refresh()

    def closePair(self, driver):
        while True:
            is_closed = False

            try:
                pairs = WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'MuiTableRow-root')))[1:]
            except TimeoutException:
                pairs = []
            
            for pair in pairs:
                is_closed = False
                cells = pair.find_elements(By.CLASS_NAME, 'MuiTableCell-root')

                name = cells[1].find_element(By.CLASS_NAME, 'two-row-cell').find_element(By.TAG_NAME, 'div').text.replace(' / ', '/')

                for pair_name in self.close_list_buff:
                    if name == pair_name:
                        print('Closing', name)
                        # cells[-1].find_elements(By.TAG_NAME, 'button')[-1].click()
                        # WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//button[@class="aQXLoSia4k1esjIDAFwW zhVSsYrxjm8vd8ihxSUs MuiButton-root MuiButton-text MuiButton-textPrimary MuiButton-sizeMedium MuiButton-textSizeMedium MuiButtonBase-root  css-pev4aq"]'))).click()
                        is_passed = False

                        while not is_passed:
                            try:
                                driver.execute_script("arguments[0].click();", cells[-1].find_elements(By.TAG_NAME, 'button')[-1])
                                driver.execute_script("arguments[0].click();", WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//button[@class="aQXLoSia4k1esjIDAFwW zhVSsYrxjm8vd8ihxSUs MuiButton-root MuiButton-text MuiButton-textPrimary MuiButton-sizeMedium MuiButton-textSizeMedium MuiButtonBase-root  css-pev4aq"]'))))
                                is_passed = True
                            except TimeoutException:
                                is_passed = False

                        # try:
                        close_options = WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.XPATH, '//li[@class="yh3uTjDDJTvbuAZD9i_M jj5mPys2QhRB6omDdQP4 MuiMenuItem-root MuiMenuItem-gutters MuiButtonBase-root css-17cm1p2"]')))

                        if len(close_options) > 1:
                            driver.execute_script("arguments[0].click();", close_options[1])
                            # close_options[1].click()
                        
                        if len(close_options) == 1:
                            driver.execute_script("arguments[0].click();", close_options[0])
                            # close_options[0].click()
                            # driver.execute_script("arguments[0].click();", WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.XPATH, '//li[@class="yh3uTjDDJTvbuAZD9i_M jj5mPys2QhRB6omDdQP4 MuiMenuItem-root MuiMenuItem-gutters MuiButtonBase-root css-17cm1p2"]')))[1])
                        # except IndexError:
                        #     driver.execute_script("arguments[0].click();", WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.XPATH, '//li[@class="yh3uTjDDJTvbuAZD9i_M jj5mPys2QhRB6omDdQP4 MuiMenuItem-root MuiMenuItem-gutters MuiButtonBase-root css-17cm1p2"]')))[0])
                        
                        confirm = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//button[@data-test="bot-preview-confirm-button"]')))
                        # confirm.click()
                        driver.execute_script("arguments[0].click();", confirm)

                        is_closed = True
                        
                        try:
                            WebDriverWait(driver, 300).until(lambda s: s.find_element(By.XPATH, '//div[@class="vfOtxORs1c0oUEoFMKdP"]').text == 'Your bot has been stopped')
                        except:
                            driver.refresh()

                        break

                if is_closed:
                    break

            if is_closed:
                continue
            else:
                break

    def extract(self, driver):
        pairs = driver.find_elements(By.CLASS_NAME, 'MuiTableRow-root')[1:]

        if len(pairs) > 0:
            pair_list = []
            change_list = []

            for pair in pairs:
                cells = pair.find_elements(By.CLASS_NAME, 'MuiTableCell-root')

                try:
                    name = cells[1].find_element(By.CLASS_NAME, 'two-row-cell').find_element(By.TAG_NAME, 'div').text.replace(' / ', '/')
                except NoSuchElementException:
                    self.progress.emit([], [], True)
                    self.progress.emit(["No pairs"], [0], False)

                    break
                
                try:
                    change = float(cells[3].text[:-1])
                except:
                    change = 0

                pair_list.append(name)
                change_list.append(change)

            self.progress.emit(pair_list, change_list, True)
        else:
            self.progress.emit([], [], True)
            self.progress.emit(["No pairs"], [0], False)

class Ui(QtWidgets.QMainWindow):
    def __init__(self):
        super(Ui, self).__init__()
        uic.loadUi('main.ui', self)

        self.last_signal = {}

        self.mt_profit.setValidator(QDoubleValidator())
        self.mt_stoploss.setValidator(QDoubleValidator())
        self.mt_viewchart.clicked.connect(self.viewMTChart)
        self.mt_clearchart.clicked.connect(self.clearMTChart)
        self.singel_system_checkbox.clicked.connect(self.updateSingleSystem)

        self.statusBar.showMessage('Loading...')

        self.bot_thread = QThread(self)
        self.bot_worker = Bot(self.single_system_table, self.single_system_inv)
        self.bot_worker.moveToThread(self.bot_thread)
        self.bot_thread.started.connect(self.bot_worker.run)
        self.bot_worker.progress.connect(self.updateStatus)
        self.bot_worker.finished.connect(self.bot_thread.quit)
        self.bot_worker.finished.connect(self.bot_worker.deleteLater)
        self.bot_thread.finished.connect(self.bot_thread.deleteLater)
        self.bot_thread.start()

        self.webhook_thread = QThread(self)
        self.webhook_worker = Webhook()
        self.webhook_worker.moveToThread(self.webhook_thread)
        self.webhook_thread.started.connect(self.webhook_worker.run)
        self.webhook_worker.progress.connect(self.updateSignal)
        self.webhook_worker.finished.connect(self.webhook_thread.quit)
        self.webhook_worker.finished.connect(self.webhook_worker.deleteLater)
        self.webhook_thread.finished.connect(self.webhook_thread.deleteLater)
        self.webhook_thread.start()

        self.pricer_thread = QThread(self)
        self.pricer_worker = Pricer(self.single_system_table)
        self.pricer_worker.moveToThread(self.pricer_thread)
        self.pricer_thread.started.connect(self.pricer_worker.run)
        self.pricer_worker.finished.connect(self.pricer_thread.quit)
        self.pricer_worker.finished.connect(self.pricer_worker.deleteLater)
        self.pricer_thread.finished.connect(self.pricer_thread.deleteLater)
        self.pricer_thread.start()

        self.chart_logger_thread = QThread(self)
        self.chart_logger_worker = ChartLogger(self.st_table)
        self.chart_logger_worker.moveToThread(self.chart_logger_thread)
        self.chart_logger_thread.started.connect(self.chart_logger_worker.run)
        self.chart_logger_thread.start()

        self.show()

    def updateSingleSystem(self, enabled):
        global is_single_sytem_activated
        is_single_sytem_activated = enabled

        self.single_system_inv_text.setEnabled(enabled)
        self.single_system_inv.setEnabled(enabled)
        self.single_system_table.setEnabled(enabled)

        self.last_signal = {}

    def keyPressEvent(self, event):
        widget = QtWidgets.QApplication.focusWidget()
        
        if widget.objectName() == 'single_system_table' or widget.objectName() == 'mt_table' or widget.objectName() == 'st_table':
            if event.key() == Qt.Key.Key_V:
                try:
                    clipboard = pd.read_clipboard(sep=r'\s+', header=None)
                except:
                    clipboard = pd.DataFrame()
                clipboard.fillna('', inplace=True)

                clip_rows = len(clipboard)
                clip_cols = len(clipboard.columns)

                for row_index in range(clip_rows):
                    for col_index in range(clip_cols):
                        try:
                            widget.setItem(widget.selectedIndexes()[0].row()+row_index, widget.selectedIndexes()[0].column()+col_index, QtWidgets.QTableWidgetItem(str(clipboard.iat[row_index, col_index])))
                        except:
                            pass
            elif event.key() == Qt.Key.Key_Delete:
                for index in widget.selectedIndexes():
                    widget.setItem(index.row(), index.column(), QtWidgets.QTableWidgetItem(''))

    def updateStatus(self, message, value, state):
        if not state:
            self.statusBar.showMessage(''.join(message))

            if ''.join(message) == 'Loading finished!':
                self.setEnabled(True)
        else:
            status = []
            
            for index in range(len(message)):
                status.append(f'{message[index]}:{value[index]}')

            self.statusBar.showMessage(' '.join(status))

            self.updateMT(message, value)
            self.updateST(message, value)

    def updateSignal(self, signal: dict):
        keys = signal.keys()
        open_list = []
        close_list = []

        if len(keys) > 0:
            for key in keys:
                if key not in self.last_signal.keys():
                    if signal[key]['signal'] == 'Buy':
                        close_list.append(key.replace('USDT', '3S/USDT'))
                        open_list.append(key.replace('USDT', '3L/USDT'))
                    elif signal[key]['signal'] == 'Sell':
                        close_list.append(key.replace('USDT', '3L/USDT'))
                        open_list.append(key.replace('USDT', '3S/USDT'))

                elif self.last_signal[key]['signal'] != signal[key]['signal']:
                    if signal[key]['signal'] == 'Buy':
                        close_list.append(key.replace('USDT', '3S/USDT'))
                        open_list.append(key.replace('USDT', '3L/USDT'))
                    elif signal[key]['signal'] == 'Sell':
                        close_list.append(key.replace('USDT', '3L/USDT'))
                        open_list.append(key.replace('USDT', '3S/USDT'))

                self.last_signal[key] = signal[key]

            # Search single system table
            row_count = self.single_system_table.rowCount()

            for pair in open_list:
                if '3L/' in pair:
                    buy = True
                    sell = False
                elif '3S/' in pair:
                    buy = False
                    sell = True

                pair_for_single_table = pair.replace('3L/', '').replace('3S/', '')

                for row_index in range(row_count):
                    if self.single_system_table.item(row_index, 0):
                        if self.single_system_table.item(row_index, 0).text() == pair_for_single_table:
                            # Add pair to st_table on 1st tab
                            st_table_row_count = self.st_table.rowCount()
                            is_exist_in_st = False

                            for st_row_index in range(st_table_row_count):
                                item = self.st_table.item(st_row_index, 0)

                                if item:
                                    if item.text() == pair:
                                        is_exist_in_st = True
                                        break

                            if not is_exist_in_st:
                                for st_row_index in range(st_table_row_count):
                                    item = self.st_table.item(st_row_index, 0)

                                    if item and item.text() == "":
                                        self.st_table.setItem(st_row_index, 0, QtWidgets.QTableWidgetItem(pair))

                                        tp = self.single_system_table.item(row_index, 9).text()
                                        sl = self.single_system_table.item(row_index, 10).text()

                                        self.st_table.setItem(st_row_index, 2, QtWidgets.QTableWidgetItem(tp))
                                        self.st_table.setItem(st_row_index, 3, QtWidgets.QTableWidgetItem(sl))

                                        current_datetime = datetime.now()
                                        current_date_time = current_datetime.strftime('%Y-%m-%d %H:%M:%S')

                                        self.st_table.setItem(st_row_index, 6, QtWidgets.QTableWidgetItem(current_date_time))

                                        break

                                    if not item:
                                        self.st_table.setItem(st_row_index, 0, QtWidgets.QTableWidgetItem(pair))

                                        tp = self.single_system_table.item(row_index, 9).text()
                                        sl = self.single_system_table.item(row_index, 10).text()

                                        self.st_table.setItem(st_row_index, 2, QtWidgets.QTableWidgetItem(tp))
                                        self.st_table.setItem(st_row_index, 3, QtWidgets.QTableWidgetItem(sl))

                                        current_datetime = datetime.now()
                                        current_date_time = current_datetime.strftime('%Y-%m-%d %H:%M:%S')

                                        self.st_table.setItem(st_row_index, 6, QtWidgets.QTableWidgetItem(current_date_time))

                                        break

                            # Update single system table state
                            if buy and not sell:
                                self.single_system_table.item(row_index, 1).setBackground(QColor(0, 255, 0))
                                self.single_system_table.item(row_index, 2).setBackground(QColor(255, 255, 255))
                            elif not buy and sell:
                                self.single_system_table.item(row_index, 1).setBackground(QColor(255, 255, 255))
                                self.single_system_table.item(row_index, 2).setBackground(QColor(255, 0, 0))

            for pair in close_list:
                st_table_row_count = self.st_table.rowCount()

                for row_index in range(st_table_row_count):
                    item = self.st_table.item(row_index, 0)

                    if item:
                        if item.text() == pair:
                            log_data = [pair.replace('/', ''), 'single']

                            current_datetime = datetime.now()
                            current_date_time = current_datetime.strftime('%Y-%m-%d %H:%M:%S')

                            open_date_time = self.st_table.item(row_index, 6).text()

                            change = self.st_table.item(row_index, 1).text()

                            log_data.append(open_date_time)
                            log_data.append(current_date_time)
                            log_data.append(change)
                            log_data.append('TV Signal')

                            ws.append(log_data)
                            wb.save('Closed Trades.xlsx')

            self.bot_worker.setClose(close_list)
            self.bot_worker.setOpen(open_list)

    def clearMTChart(self):
        global track
        track['MT'] = []

    def viewMTChart(self, is_export=False):
        graphs = {}
        tp = []
        sl = []
        timestamps = []

        if len(track['MT']) > 0:
            for key in track['MT'][-1].keys():
                if key != "Exist" and key != 'timestamp':
                    graphs[key] = []

            for index in range(len(track['MT'])):
                if track['MT'][index]['Exist']:
                    timestamps.append(track['MT'][index]['timestamp'])
                    tp.append(float(self.mt_profit.text()))
                    sl.append(float(self.mt_stoploss.text()))
                    for key in graphs.keys():
                        try:
                            graphs[key].append(track['MT'][index][key])
                        except:
                            graphs[key].append(0)

            plt.figure('Multiple chart')

            datenums = md.date2num(timestamps)

            ax = plt.gca()
            xfmt = md.DateFormatter('%m-%d %H:%M:%S')
            # plt.xlim(datetime(bot_start_date.year, bot_start_date.month, bot_start_date.day, 0, 0, 0), datetime(datetime.today().year, datetime.today().month, datetime.today().day, 23, 59, 59))
            ax.xaxis.set_major_formatter(xfmt)
            # ax.xaxis.set_major_locator(SecondLocator(bysecond=range(60), interval=10, tz=None))
            ax.yaxis.set_major_locator(MaxNLocator(integer=True))
            
            plt.xticks(rotation=25)

            for key in graphs.keys():
                if key != "collective":
                    plt.plot(datenums, graphs[key], label=key)

            plt.plot(datenums, graphs['collective'], label='collective', linewidth=4)
            plt.plot(datenums, tp, label='TP', linewidth=4)
            plt.plot(datenums, sl, label='SL', linewidth=4)

            plt.legend(loc='lower right')
            plt.grid()

            if is_export:
                fname = datetime.now().strftime('%Y_%m_%d_%H_%M_%S') + '.png'
                plt.savefig(os.path.join(chart_dir, fname))
            else:
                plt.show()
            
    def viewSTChart(self, is_export=False, r_index=0):
        graph = []
        tp = []
        sl = []
        timestamps = []

        if not is_export:
            button = QtWidgets.QApplication.focusWidget()
            index = self.st_table.indexAt(button.pos())

            if index.isValid():
                row_index = index.row()
        else:
            row_index = r_index

        pair = self.st_table.item(row_index, 0).text()

        for index in range(len(track['ST'])):
            if pair in track['ST'][index].keys():
                try:
                    tp.append(float(self.st_table.item(row_index, 2).text()))
                except:
                    tp.append(0)

                try:
                    sl.append(float(self.st_table.item(row_index, 3).text()))
                except:
                    sl.append(0)

                timestamps.append(track['ST'][index]['timestamp'])
                graph.append(track['ST'][index][pair])

        if len(graph) > 0:
            plt.figure(pair)

            datenums = md.date2num(timestamps)

            ax = plt.gca()
            xfmt = md.DateFormatter('%m-%d %H:%M:%S')
            ax.xaxis.set_major_formatter(xfmt)
            # plt.xlim(datetime(bot_start_date.year, bot_start_date.month, bot_start_date.day, 0, 0, 0), datetime(datetime.today().year, datetime.today().month, datetime.today().day, 23, 59, 59))
            # ax.xaxis.set_major_locator(SecondLocator(bysecond=range(60), interval=10, tz=None))
            ax.yaxis.set_major_locator(MaxNLocator(integer=True))

            plt.xticks(rotation=25)

            plt.plot(datenums, graph, label=pair)
            plt.plot(datenums, tp, label='TP', linewidth=4)
            plt.plot(datenums, sl, label='SL', linewidth=4)
            plt.legend(loc='lower right')
            plt.grid()

            if is_export:
                fname = datetime.now().strftime('%Y_%m_%d_%H_%M_%S') + '.png'
                plt.savefig(os.path.join(chart_dir, fname))
            else:
                plt.show()

    def clearSTChart(self, is_export=False, r_index=0):
        if not is_export:
            button = QtWidgets.QApplication.focusWidget()
            index = self.st_table.indexAt(button.pos())

            if index.isValid():
                row_index = index.row()
        else:
            row_index = r_index

        pair = self.st_table.item(row_index, 0).text()

        for step in track['ST']:
            if pair in step.keys():
                step.pop(pair, None)

    def updateMT(self, pairs, changes):
        row_count = self.mt_table.rowCount()
        existing_changes = []
        existing_pairs = []
        collective = 0
        global track
        data = {
            'Exist': False
        }
        
        for row_index in range(row_count):
            item = self.mt_table.item(row_index, 0)
            
            if item:
                if item.text() != "" and item.text() in pairs:
                    self.mt_table.setItem(row_index, 1, QtWidgets.QTableWidgetItem('{:.2f}'.format(changes[pairs.index(item.text())])))
                    data[item.text()] = changes[pairs.index(item.text())]
                    data['Exist'] = True
                    existing_pairs.append(item.text())
                    existing_changes.append(changes[pairs.index(item.text())])
                else:
                    self.mt_table.setItem(row_index, 1, QtWidgets.QTableWidgetItem(''))

        if len(existing_changes) > 0:
            collective = sum(existing_changes) / len(existing_changes)

        self.mt_collective.setText("{:.2f}".format(collective))
        data['collective'] = collective
        data['timestamp'] = datetime.now()

        track['MT'].append(data)

        if self.mt_stoploss.text() != "" and self.mt_profit.text() != "" and self.mt_stoploss.text() != "." and self.mt_profit.text() != ".":
            sl = float(self.mt_stoploss.text())
            tp = float(self.mt_profit.text())

            if collective <= sl:
                alarm()
                for pair in existing_pairs:
                    for row_index in range(self.mt_table.rowCount()):
                        if self.mt_table.item(row_index, 0):
                            if self.mt_table.item(row_index, 0).text() == pair:
                                self.viewMTChart(is_export=True)
                                log_data = [pair, "multiple"]

                                current_datetime = datetime.now()

                                current_date_time = current_datetime.strftime('%Y-%m-%d %H:%M:%S')
                                open_date_time = self.mt_table.item(row_index, 2).text()

                                log_data.append(open_date_time)
                                log_data.append(current_date_time)
                                log_data.append(float(self.mt_table.item(row_index, 1).text()))
                                log_data.append("Hit SL")
                                log_data.append(tp)
                                log_data.append(sl)
                                log_data.append(collective)

                                ws.append(log_data)
                                wb.save('Closed Trades.xlsx')

                                for col_index in range(2):
                                    self.mt_table.setItem(row_index, col_index, QtWidgets.QTableWidgetItem(''))

                self.bot_worker.setClose(existing_pairs)
            elif collective >= tp:
                alarm()
                for pair in existing_pairs:
                    for row_index in range(self.mt_table.rowCount()):
                        if self.mt_table.item(row_index, 0):
                            if self.mt_table.item(row_index, 0).text() == pair:
                                self.viewMTChart(is_export=True)
                                log_data = [pair, "multiple"]

                                current_datetime = datetime.now()

                                current_date_time = current_datetime.strftime('%Y-%m-%d %H:%M:%S')
                                open_date_time = self.mt_table.item(row_index, 2).text()

                                log_data.append(open_date_time)
                                log_data.append(current_date_time)
                                log_data.append(float(self.mt_table.item(row_index, 1).text()))
                                log_data.append("Hit TP")
                                log_data.append(tp)
                                log_data.append(sl)
                                log_data.append(collective)

                                ws.append(log_data)
                                wb.save('Closed Trades.xlsx')

                                for col_index in range(2):
                                    self.mt_table.setItem(row_index, col_index, QtWidgets.QTableWidgetItem(''))

                self.bot_worker.setClose(existing_pairs)

    def updateST(self, pairs, changes):
        row_count = self.st_table.rowCount()
        delete_pairs = []
        global track
        data = {
            'Exist': False
        }
        
        for row_index in range(row_count):
            item = self.st_table.item(row_index, 0)
            
            if item:
                if item.text() != "" and item.text() in pairs:
                    chart_btn = QtWidgets.QPushButton("Show")
                    chart_btn.clicked.connect(self.viewSTChart)

                    clear_btn = QtWidgets.QPushButton("Clear")
                    clear_btn.clicked.connect(self.clearSTChart)

                    self.st_table.setItem(row_index, 1, QtWidgets.QTableWidgetItem('{:.2f}'.format(changes[pairs.index(item.text())])))
                    self.st_table.setCellWidget(row_index, 4, chart_btn)
                    self.st_table.setCellWidget(row_index, 5, clear_btn)

                    data[item.text()] = changes[pairs.index(item.text())]
                    data['Exist'] = True
                    change = changes[pairs.index(item.text())]
                    tp = self.st_table.item(row_index, 2)
                    sl = self.st_table.item(row_index, 3)

                    if tp and sl:
                        if tp.text() != "" and tp.text() != "." and sl.text() != "" and sl.text() != ".":
                            if change >= float(tp.text()) or change <= float(sl.text()):
                                alarm()
                                delete_pairs.append(item.text())
                else:
                    for col_index in range(4):
                        self.st_table.setItem(row_index, col_index, None)

                    self.st_table.setItem(row_index, 6, None)
                    
                    self.st_table.setCellWidget(row_index, 4, None)
                    self.st_table.setCellWidget(row_index, 5, None)

        data['timestamp'] = datetime.now()
        track['ST'].append(data)

        if len(delete_pairs) > 0:
            for pair in delete_pairs:
                for row_index in range(self.st_table.rowCount()):
                    if self.st_table.item(row_index, 0):
                        if self.st_table.item(row_index, 0).text() == pair:
                            self.viewSTChart(is_export=True, r_index=row_index)
                            log_data = [pair, "single"]

                            current_datetime = datetime.now()

                            current_date_time = current_datetime.strftime('%Y-%m-%d %H:%M:%S')
                            open_date_time = self.st_table.item(row_index, 6).text()

                            log_data.append(open_date_time)
                            log_data.append(current_date_time)

                            change = float(self.st_table.item(row_index, 1).text())
                            tp = float(self.st_table.item(row_index, 2).text())
                            sl = float(self.st_table.item(row_index, 3).text())

                            log_data.append(change)

                            if change >= tp:
                                log_data.append("Hit TP")
                                log_data.append(tp)
                                log_data.append(sl)
                            elif change <= sl:
                                log_data.append("Hit SL")
                                log_data.append(tp)
                                log_data.append(sl)

                            ws.append(log_data)
                            wb.save('Closed Trades.xlsx')

                            for col_index in range(4):
                                self.st_table.setItem(row_index, col_index, QtWidgets.QTableWidgetItem(''))

                            self.st_table.setCellWidget(row_index, 4, None)
                            self.st_table.setCellWidget(row_index, 5, None)

            self.bot_worker.setClose(delete_pairs)

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    window = Ui()
    app.exec()

    wb.close()
