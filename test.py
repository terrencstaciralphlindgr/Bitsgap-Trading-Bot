from openpyxl import load_workbook

wb = load_workbook('Chart Data Log.xlsx')
ws = wb.worksheets[0]

column_count = ws.max_column
pairs = []

for index in range(3, column_count+1):
    pairs.append(ws.cell(1, index).value)

print(pairs)